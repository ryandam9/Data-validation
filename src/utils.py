import os
import textwrap
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles.borders import Border, Side
from tabulate import tabulate


def write_to_excel_file(list1, list2):
    """
    Writes the input M x N matrices to an Excel file.

    Assumption is that, both the lists are of the same size (i.e, each have
    M rows, and there are N cells in each row).

    :param list1: M x N matrix
    :param list2: M x N matrix

    :return: None
    """
    no_cells_in_each_row_in_list1 = set([len(row) for row in list1])
    no_cells_in_each_row_in_list2 = set([len(row) for row in list2])

    if (
        len(list1) != len(list2)
        or no_cells_in_each_row_in_list1 != no_cells_in_each_row_in_list2
    ):
        print("Input lists are not of the same size.")

        print(
            tabulate(
                list1[1:],
                headers=list1[0],
                tablefmt="fancy_grid",
            )
        )

        print(
            tabulate(
                list2[1:],
                headers=list2[0],
                tablefmt="fancy_grid",
            )
        )

        return

    wb = openpyxl.Workbook()
    sheet = wb["Sheet"]  # Default sheet name is 'Sheet'
    sheet.title = "structure_comparison"
    sheet.sheet_properties.tabColor = "1072BA"

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    current_time = (
        datetime.now().strftime("%Y_%m_%d %H:%M").replace(" ", "_").replace(":", "_")
    )

    target_file = (
        "../table_structure_validation/structure_comparison"
        + "_"
        + current_time
        + ".xlsx"
    )

    for i in range(len(list1)):
        for j in range(len(list1[i])):
            sheet.cell(row=i + 1, column=j + 1).value = list1[i][j]
            sheet.cell(row=i + 1, column=j + 1).border = thin_border

        k = len(list1[i]) + 1

        for j in range(len(list2[i])):
            sheet.cell(row=i + 1, column=k + j + 1).value = list2[i][j]
            sheet.cell(row=i + 1, column=k + j + 1).border = thin_border

    wb.save(target_file)

    print(f"-> Data written to excel file: {target_file}")


def print_messages(messages, headers):
    """
    Prints the messages in the input list.

    :param messages: List of Lists
    :param headers: List of strings

    :return: None
    """
    wrapped_messages = []

    for message in messages:
        wrapped_message = "\n".join(
            textwrap.wrap(message[0], width=180, replace_whitespace=False)
        )

        wrapped_messages.append([wrapped_message])

    print(
        tabulate(
            wrapped_messages,
            headers=headers,
            tablefmt="fancy_grid",
        )
    )


def get_tables_to_validate() -> list:
    """
    This function reads the "tables.txt" file in the project root folder and returns 
    a list of tables to validate.

    :return: A list of tables to validate. Each table is a map with keys:
        'schema' and 'table'
    """
    tables = []

    try:
        file_full_path = os.path.join(get_project_root(), 'tables.txt')
        print(f"-> Reading {file_full_path}")

        with open(file_full_path, "r") as f:
            for line in f:
                schema, table = line.split(",")[0], line.split(",")[1]
                schema = schema.strip().upper()
                table = table.strip().upper()

                tables.append({"schema": schema, "table": table})

        tables.sort(key=lambda x: x["schema"] + x["table"])
        return tables
    except Exception as err:
        print(f"-> Error: When trying to read {file_full_path}")
        print(f"-> Error: {err}")
        return []


def get_current_time():
    return (
        datetime.now().strftime("%Y_%m_%d %H:%M").replace(" ", "_").replace(":", "_")
    )


def get_project_root():
    return Path(__file__).parent.parent
